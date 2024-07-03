from books import airbnb, booking, web
import openpyxl
from openpyxl import Workbook

class read_excel:
    def __init__(self, file_in):
        self.file_in = file_in
        self.platforms = {
        "airbnb": airbnb,
        "booking": booking,
        "mara": web
        }

    def read_excel(self):
        wb = openpyxl.load_workbook(self.file_in)
        hoja = wb.active
        last = hoja.max_row
        
        reservas = []

        for i in range(1,last):
            fecha = str(hoja.cell(i+1,5).value)
            factura = "MBA24/" + str(hoja.cell(i+1,4).value)
            nombre = hoja.cell(i+1,14).value
            NIF = hoja.cell(i+1,15).value
            Base_1 = hoja.cell(i+1,18).value
            Cuota_1 = hoja.cell(i+1,19).value
            total = hoja.cell(i+1,20).value
            domicilio = hoja.cell(i+1,11).value
            cod_postal = hoja.cell(i+1,12).value
            pais = hoja.cell(i+1,10).value
            CL = 2
            observaciones = hoja.cell(i+1,26).value
            tipo_Sii = 1

            #Intenta crear las reservas. Si hay carácteres fantasmas continua a la siguiente iteración
            try:
                obs=observaciones.lower()
                #Comprueba en la lista de plataformas si coincide con las observaciones
                #Esto se puede optimizar??
                for i in self.platforms:
                    if i in obs:
                        #Hace "plataformas[i]" porque i es un string del diccionario e platforms[i] es el objeto
                        reservas.append(self.platforms[i](fecha, factura, nombre, NIF, Base_1, Cuota_1, total, domicilio, cod_postal, pais, CL, observaciones, 430000, tipo_Sii))
            except AttributeError:
                continue

        wb.close()
        return reservas
