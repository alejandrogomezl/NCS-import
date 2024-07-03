from books import airbnb, booking, web
from openpyxl import Workbook
import openpyxl
import os
import shutil
import pandas as pd

class write_excel:

    def __init__(self, reservas):
        self.reservas = reservas

    def empty_out(self):
        try:
            shutil.rmtree("out")
        except:
            pass
        os.mkdir("out")
    
    def conv(self, file):
        df = pd.read_excel(file, engine='openpyxl')
        file_new = file[:-5] + ".xls"
        with pd.ExcelWriter(file_new, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        os.remove(file)
        print(f"Archivo {file_new} convertido a .xls")

    def write (self):
        self.empty_out()
        for i in self.reservas:
            clase = str(type(i).__name__)               
            file_out = "out/" + clase + ".xlsx"

            try:
                wb = openpyxl.load_workbook(file_out)
            except FileNotFoundError:
                wb = Workbook()

            hoja = wb.active
            last = hoja.max_row

            hoja.cell(last+1,1).value = i.fecha
            hoja.cell(last+1,2).value = i.factura
            hoja.cell(last+1,3).value = i.nombre
            hoja.cell(last+1,4).value = i.NIF
            hoja.cell(last+1,5).value = i.Base_1
            hoja.cell(last+1,7).value = i.Cuota_1
            hoja.cell(last+1,8).value = i.total
            hoja.cell(last+1,9).value = i.domicilio
            hoja.cell(last+1,10).value = i.cod_postal
            hoja.cell(last+1,11).value = i.pais
            hoja.cell(last+1,12).value = i.CL
            hoja.cell(last+1,14).value = i.observaciones
            hoja.cell(last+1,15).value = i.cuenta_contable
            hoja.cell(last+1,16).value = i.tipo_Sii
            last += 1
            wb.save(file_out)
            wb.close()
            print(f"Factura {i.factura} añadida a archivo {file_out}")
        
        for element in os.listdir("out"):
            self.conv("out/"+element)
